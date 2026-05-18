from __future__ import annotations

import re
from pathlib import Path
from zipfile import BadZipFile


def extract_text(path: Path) -> tuple[str, str]:
    suffix = path.suffix.lower()
    try:
        if suffix == ".pdf":
            return _extract_pdf(path), "ok"
        if suffix in {".docx", ".docm"}:
            return _extract_docx(path), "ok"
        if suffix in {".xlsx", ".xlsm"}:
            return _extract_xlsx(path), "ok"
        if suffix in {".txt", ".csv", ".md", ".html", ".htm"}:
            return path.read_text(encoding="utf-8", errors="replace"), "ok"
        return "", f"unsupported file type: {suffix or 'unknown'}"
    except Exception as exc:  # keep sync resilient per document
        return "", f"failed: {exc}"


def _extract_pdf(path: Path) -> str:
    from pypdf import PdfReader

    reader = PdfReader(str(path))
    pages = [page.extract_text() or "" for page in reader.pages]
    return "\n\n".join(pages).strip()


def _extract_docx(path: Path) -> str:
    from docx import Document

    doc = Document(str(path))
    return "\n".join(paragraph.text for paragraph in doc.paragraphs if paragraph.text).strip()


def _extract_xlsx(path: Path) -> str:
    from openpyxl import load_workbook

    wb = load_workbook(str(path), read_only=True, data_only=True)
    lines: list[str] = []
    for sheet in wb.worksheets:
        lines.append(f"# Sheet: {sheet.title}")
        for row in sheet.iter_rows(values_only=True):
            values = [str(cell) for cell in row if cell not in (None, "")]
            if values:
                lines.append(" | ".join(values))
    return "\n".join(lines).strip()


def extract_signals(text: str) -> dict[str, list[str]]:
    sample = text[:20000]
    dates = sorted(set(re.findall(r"\b(?:\d{1,2}/\d{1,2}/\d{2,4}|[A-Z][a-z]+\s+\d{1,2},\s+\d{4})\b", sample)))[:20]
    dollars = sorted(set(re.findall(r"\$\s?\d[\d,]*(?:\.\d{2})?", sample)))[:20]
    people_depts = sorted(set(re.findall(r"\b(?:Business Office|Superintendent|Board of Education|Transportation|Technology|Special Education|Athletics|Facilities|Curriculum|Human Resources|Personnel)\b", sample, re.I)))[:20]
    words = re.findall(r"\b[A-Za-z][A-Za-z-]{4,}\b", sample.lower())
    stop = {"about", "after", "again", "board", "district", "meeting", "school", "their", "there", "these", "which", "would", "shall", "report", "agenda", "document"}
    counts: dict[str, int] = {}
    for word in words:
        if word not in stop:
            counts[word] = counts.get(word, 0) + 1
    keywords = [word for word, _ in sorted(counts.items(), key=lambda item: item[1], reverse=True)[:12]]
    return {"dates": dates, "dollars": dollars, "people_departments": people_depts, "keywords": keywords}
